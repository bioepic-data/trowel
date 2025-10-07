"""API wrapper for ESS-DIVE."""

# See https://api.ess-dive.lbl.gov/#/Dataset/getDataset

# TODO: fix frequency counting as some variables are definitely
# getting counted multiple times.
# That is, counts are by file, not by dataset.

from io import StringIO
import sys
import string
import os
import tempfile
import re
from typing import Tuple, List, Optional
import csv
import logging
import polars as pl
import requests
import xml.etree.ElementTree as ET
from tqdm import tqdm
import openpyxl
import xlrd

from trowel.utils.string_utils import clean_unicode_chars, extract_units, clean_punctuation

BASE_URL = "https://api.ess-dive.lbl.gov"

ENDPOINT = "packages"

USER_HEADERS = {
    "user_agent": "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:77.0) Gecko/20100101 Firefox/77.0",
    "content-type": "application/json",
    "Range": "bytes=0-1000",
}

PARSIBLE_ENCODINGS = ["text/csv",
                      "text/plain",
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                      "https://eml.ecoinformatics.org/eml-2.2.0",
                      ]

PARSIBLE_EXTENSIONS = [
    ".csv",
    ".txt",
    ".tsv",
    ".xlsx",
    ".xls",
    ".xml",
]

logger = logging.getLogger(__name__)


def sanitize_tsv_field(value) -> str:
    """Normalize a value destined for TSV output.

    Replaces any newlines / carriage returns / tabs with a single space,
    collapses consecutive whitespace, and strips leading/trailing spaces.
    Non-string values are coerced to string. None becomes empty string.
    """
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    # Replace problematic control characters with space
    value = value.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    # Collapse multiple whitespace to a single space
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def _norm_header_key(h: str) -> str:
    """Normalize a CSV header key for matching regardless of case/spacing/punct."""
    if not isinstance(h, str):
        h = str(h)
    return re.sub(r"[^a-z0-9]", "", h.lower())


def parse_flmd_file(content: str) -> dict:
    """Parse an FLMD (File Level Metadata) file and return a mapping of filename -> description.

    Args:
        content: The FLMD file content as a string

    Returns:
        Dictionary mapping filename to file description
    """
    file_descriptions = {}
    try:
        reader = csv.DictReader(StringIO(content))
        if not reader.fieldnames:
            return file_descriptions

        # Find the columns for file name and description (case insensitive)
        filename_col = None
        description_col = None

        for field in reader.fieldnames:
            norm_field = _norm_header_key(field)
            if norm_field in ["filename", "file_name", "name"]:
                filename_col = field
            elif norm_field in ["filedescription", "file_description", "description"]:
                description_col = field

        if not filename_col or not description_col:
            logger.debug(
                f"Could not find required columns in FLMD file. Available columns: {reader.fieldnames}")
            return file_descriptions

        # Parse the rows
        for row in reader:
            filename = row.get(filename_col, "").strip()
            description = row.get(description_col, "").strip()

            if filename and description:
                file_descriptions[filename] = sanitize_tsv_field(description)

    except Exception as e:
        logger.debug(f"Error parsing FLMD file: {e}")

    return file_descriptions


def append_dd_content_to_file(content: str, dataset_id: str, source_filename: str, out_path: str) -> int:
    """Append rows from a data dictionary CSV (comma-delimited) to a compiled TSV file.

    The compiled TSV columns are: dataset_id, source_filename, and the canonical DD columns.
    Returns number of rows appended.
    """
    canonical_columns = [
        "Column_or_Row_Name",
        "Unit",
        "Definition",
        "Column_or_Row_Long_Name",
        "Data_Type",
        "Term_Type",
        "Missing_Value_Code",
    ]
    canonical_norm_map = {
        "columnorrowname": "Column_or_Row_Name",
        "unit": "Unit",
        "definition": "Definition",
        "columnorrowlongname": "Column_or_Row_Long_Name",
        "datatype": "Data_Type",
        "termtype": "Term_Type",
        "missingvaluecode": "Missing_Value_Code",
    }

    reader = csv.DictReader(StringIO(content))
    if not reader.fieldnames:
        return 0

    # Map canonical column -> actual header in this file (if present)
    source_map = {}
    for field in reader.fieldnames:
        norm = _norm_header_key(field)
        if norm in canonical_norm_map and canonical_norm_map[norm] not in source_map:
            source_map[canonical_norm_map[norm]] = field

    rows_written = 0
    # Row value markers to exclude (normalized)
    metadata_markers = {"columnorrowname", "filedescription",
                        "missingvaluecodes", "missingvaluecode"}
    with open(out_path, "a", newline="") as f:
        writer = csv.writer(f, delimiter="\t", lineterminator="\n")
        for row in reader:
            # Determine if this row is a metadata row we should skip
            name_src = source_map.get("Column_or_Row_Name")
            name_val = row.get(name_src, "") if name_src else ""
            name_norm = _norm_header_key(
                sanitize_tsv_field(name_val)) if name_val else ""
            if name_norm in metadata_markers:
                continue

            # Apply variable name validation to the Column_or_Row_Name
            if name_val:
                # Use normalize_variables to check if this is a valid variable name
                normalized_names = normalize_variables([name_val])
                if not normalized_names:  # If normalization filtered it out, skip this row
                    continue

            # Collect canonical values
            canon_vals = []
            for col in canonical_columns:
                src = source_map.get(col)
                val = row.get(src, "") if src else ""
                canon_vals.append(sanitize_tsv_field(val))

            # Skip completely empty rows across canonical fields
            if not any(canon_vals):
                continue

            out_vals = [sanitize_tsv_field(dataset_id), sanitize_tsv_field(
                source_filename)] + canon_vals
            writer.writerow(out_vals)
            rows_written += 1
    return rows_written


def get_metadata(
    identifiers: list, token: str, outpath: str = "."
) -> Tuple[str, str, str]:
    """Get metadata from ESS-DIVE for a list of identifiers.
    The identifiers should be DOIs.
    This also requires an authentication token for ESS-DIVE.

    Results are streamed to files in the specified output directory as they are received.

    Args:
        identifiers: List of DOI identifiers
        token: ESS-DIVE authentication token
        outpath: Directory to write output files (defaults to current directory)

    Returns:
        Tuple containing paths to the results, variables frequency file, and file table file
    """
    # Create output file paths
    results_path = f"{outpath}/results.tsv"
    filetable_path = f"{outpath}/filetable.tsv"

    # Initialize empty files with headers
    results_schema = pl.DataFrame(
        schema={
            "doi": pl.Utf8,
            "id": pl.Utf8,
            "name": pl.Utf8,
            "variables": pl.Utf8,
            "description": pl.Utf8,
            "site_description": pl.Utf8,
            "methods": pl.Utf8,
        }
    )
    results_schema.write_csv(results_path, separator="\t")

    files_schema = pl.DataFrame(
        schema={
            "dataset_id": pl.Utf8,
            "url": pl.Utf8,
            "name": pl.Utf8,
            "encoding": pl.Utf8,
        }
    )
    files_schema.write_csv(filetable_path, separator="\t")

    all_variables = {}  # key is variable name, value is frequency
    headers = {"Authorization": f"Bearer {token}"}

    results_found = False
    files_found = False

    # Dictionary to collect errors
    errors = {
        "no_variables": [],
        "no_site_description": [],
        "no_methods": [],
        "no_files": [],
        "authorization": [],
        "not_found": [],
        "other_errors": []
    }

    # Add tqdm progress bar
    for identifier in tqdm(identifiers, desc="Processing identifiers", unit="entry"):
        # Check on identifier format first
        if identifier.startswith("https://doi.org/"):
            # This is a full DOI, so we need to strip it down
            identifier = identifier.replace("https://doi.org/", "")
        elif identifier.startswith("doi.org/"):
            identifier = identifier.replace("doi.org/", "")

        # clean it up
        identifier = identifier.strip()

        # Check if this is a DOI anyway
        if identifier.startswith("ess-dive"):
            sys.exit(
                f"The provided identifier {identifier} does not appear to be a DOI. Please check the format."
            )
        if not identifier.startswith("doi:"):
            identifier = "doi:" + identifier

        get_packages_url = "{}/{}/{}?&isPublic=true".format(
            BASE_URL, ENDPOINT, identifier
        )
        response = requests.get(
            get_packages_url, headers=headers, verify=True, stream=True)

        if response.status_code == 200:
            # Success - but will need to restructure
            these_results = response.json()
            # Add relevant parts to the dataframe
            essdive_id = these_results["id"]
            name = sanitize_tsv_field(these_results["dataset"]["name"])
            try:
                variables = normalize_variables(
                    these_results["dataset"]["variableMeasured"]
                )
                # Sanitize variable names (defensive – normalize_variables should already do most cleaning)
                variables = [sanitize_tsv_field(v) for v in variables if v]
                for var in variables:
                    if var in all_variables:
                        all_variables[var] += 1
                    else:
                        all_variables[var] = 1
            except KeyError:
                errors["no_variables"].append(identifier)
                variables = []
            desc_text = these_results["dataset"].get("description", [])
            try:
                site_desc = these_results["dataset"]["spatialCoverage"][0][
                    "description"
                ]
            except KeyError:
                errors["no_site_description"].append(identifier)
                site_desc = ""
            try:
                methods = these_results["dataset"]["measurementTechnique"]
            except KeyError:
                errors["no_methods"].append(identifier)
                methods = []

            # Handle cases where description / methods may be strings instead of lists
            if isinstance(desc_text, str):
                desc_joined = desc_text
            else:
                # Some APIs return list/array of strings
                desc_joined = " ".join([d for d in desc_text if d])
            if isinstance(methods, str):
                methods_joined = methods
            else:
                methods_joined = " ".join([m for m in methods if m])

            # Sanitize fields for TSV safety
            description_clean = sanitize_tsv_field(desc_joined)
            site_desc_clean = sanitize_tsv_field(site_desc)
            methods_clean = sanitize_tsv_field(methods_joined)

            # Create entry for results and append to file
            entry = pl.DataFrame(
                {
                    "doi": [identifier],
                    "id": [essdive_id],
                    "name": [name],
                    "variables": ["|".join(variables)],
                    "description": [description_clean],
                    "site_description": [site_desc_clean],
                    "methods": [methods_clean],
                }
            )
            # Append to results file
            with open(results_path, "a") as f:
                entry.write_csv(f, separator="\t", include_header=False)
            results_found = True

            # See if we have file information
            if "distribution" in these_results["dataset"]:
                for raw_entry in these_results["dataset"]["distribution"]:
                    url = raw_entry["contentUrl"]
                    filename = raw_entry["name"]
                    encoding = raw_entry["encodingFormat"]
                    file_entry = pl.DataFrame(
                        {
                            "dataset_id": [essdive_id],
                            "url": [url],
                            "name": [filename],
                            "encoding": [encoding],
                        }
                    )
                    # Append to filetable
                    with open(filetable_path, "a") as f:
                        file_entry.write_csv(
                            f, separator="\t", include_header=False)
                    files_found = True
            else:
                errors["no_files"].append(identifier)
        else:
            # There was an error
            if response.status_code == 401:
                errors["authorization"].append(identifier)
                logger.error(
                    "You may need to refresh your authentication token for ESS-DIVE."
                )
                break
            elif response.status_code == 404:
                errors["not_found"].append(identifier)
            else:
                errors["other_errors"].append(
                    f"{identifier} (status code: {response.status_code})")
                break

    # Check if files have content and log errors if they don't
    if not results_found:
        logger.error(
            "No metadata results were found for the provided identifiers")

    if not files_found:
        logger.error("No files were found for the provided identifiers")

    # Write frequencies to file
    frequencies_path = f"{outpath}/frequencies.tsv"
    with open(frequencies_path, "w") as freq_file:
        for freq in all_variables:
            freq_file.write(f"{freq}\t{all_variables[freq]}\n")

    # Log all collected errors
    if any(errors.values()):
        logger.error("The following errors occurred during processing:")

        if errors["no_variables"]:
            logger.error(
                f"No variables found for: {', '.join(errors['no_variables'])}")

        if errors["no_site_description"]:
            logger.error(
                f"No site description found for: {', '.join(errors['no_site_description'])}")

        if errors["no_methods"]:
            logger.error(
                f"No methods found for: {', '.join(errors['no_methods'])}")

        if errors["no_files"]:
            logger.error(
                f"No files found for: {', '.join(errors['no_files'])}")

        if errors["authorization"]:
            logger.error(
                f"Authorization errors for: {', '.join(errors['authorization'])}")

        if errors["not_found"]:
            logger.error(
                f"Datasets not found for: {', '.join(errors['not_found'])}")

        if errors["other_errors"]:
            logger.error(f"Other errors: {', '.join(errors['other_errors'])}")

    return results_path, frequencies_path, filetable_path


def parse_excel_header(content: bytes, filename: str) -> List[str]:
    """Parse header from an Excel file.
    Also normalizes column names and filters out excessively long names (>70 chars).

    Args:
        content: The Excel file content as bytes
        filename: Name of the file (used to determine if it's .xls or .xlsx)

    Returns:
        List of normalized column names
    """
    header_names = []
    file_ext = os.path.splitext(filename.lower())[1]

    # Create a temporary file to save the Excel content
    with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as temp_file:
        temp_file.write(content)
        temp_path = temp_file.name

    try:
        if file_ext == '.xlsx':
            # Use openpyxl for .xlsx files
            wb = openpyxl.load_workbook(
                temp_path, read_only=True, data_only=True)

            # Get the first worksheet
            ws = wb.active

            # Read the first row (header)
            # type: ignore[attr-defined]
            first_row = next(ws.iter_rows(
                min_row=1, max_row=1, values_only=True))
            for cell_val in first_row:
                if cell_val and isinstance(cell_val, (str, int, float)):
                    value = str(cell_val).strip()
                    # Skip empty names and names > 70 chars
                    if value and len(value) <= 70:
                        header_names.append(value.lower().replace("_", " "))
        elif file_ext == '.xls':
            # Use xlrd for .xls files
            wb = xlrd.open_workbook(temp_path)

            # Get the first worksheet
            ws = wb.sheet_by_index(0)

            # Read the first row (header)
            for col in range(ws.ncols):
                value = ws.cell_value(0, col)
                if value:
                    value = str(value).strip()
                    # Skip empty names and names > 70 chars
                    if value and len(value) <= 70:
                        header_names.append(value.lower().replace("_", " "))
    except Exception as e:
        logger.debug(f"Error parsing Excel file {filename}: {str(e)}")
    finally:
        # Clean up the temporary file
        try:
            os.unlink(temp_path)
        except:
            pass

    return header_names


def get_file_extension(filename: str) -> str:
    """Get the lowercase extension of a file."""
    return os.path.splitext(filename.lower())[1]


def get_variable_names(filetable_path: str, results_path: Optional[str] = None, outpath: str = ".") -> str:
    """Get dataset variable names from ESS-DIVE for a list of data identifiers.
    Takes the name/path of the table, as produced by get_metadata,
    as input.

    Variable names and keywords are streamed to an output file as they are collected.
    This also includes data dictionary contents if present for a given dataset.

    Args:
        filetable_path: Path to the filetable created by get_metadata
        results_path: Path to the results file (TSV) - needed for dataset mapping
        outpath: Directory to write output file (defaults to current directory)

    Returns:
        Path to the column names output file
    """
    # Define the output file path
    variable_names_path = f"{outpath}/variable_names.tsv"

    # Initialize a compiled data dictionaries output
    data_dict_compiled_path = f"{outpath}/data_dictionaries.tsv"
    with open(data_dict_compiled_path, "w", newline="") as _dd_out:
        dd_writer = csv.writer(_dd_out, delimiter="\t", lineterminator="\n")
        dd_writer.writerow([
            "dataset_id",
            "source_filename",
            "Column_or_Row_Name",
            "Unit",
            "Definition",
            "Column_or_Row_Long_Name",
            "Data_Type",
            "Term_Type",
            "Missing_value_code",
        ])

    # Initialize the column, keyword, and data dictionary frequency dictionaries for tracking
    column_frequencies = {}
    keyword_frequencies = {}

    # Load dataset mapping if results file is provided
    dataset_mapping = {}  # maps variable -> list of (dataset_id, dataset_name)
    dataset_id_to_name = {}  # maps dataset_id -> dataset_name
    if results_path and os.path.exists(results_path):
        try:
            results_df = pl.read_csv(results_path, separator="\t")
            for row in results_df.iter_rows(named=True):
                dataset_id = row["id"]
                dataset_name = row["name"]

                # Build dataset_id to dataset_name mapping
                dataset_id_to_name[dataset_id] = dataset_name

                variables_str = row.get("variables", "")
                if variables_str:
                    # Split pipe-delimited variables
                    variables = [v.strip()
                                 for v in variables_str.split("|") if v.strip()]
                    for var in variables:
                        if var not in dataset_mapping:
                            dataset_mapping[var] = []
                        dataset_mapping[var].append((dataset_id, dataset_name))
        except Exception as e:
            logger.warning(
                f"Could not load results file for dataset mapping: {e}")
            dataset_mapping = {}
            dataset_id_to_name = {}
    data_dict_frequencies = {}
    # Track definitions and units for variables from data dictionaries
    variable_definitions = {}
    variable_units = {}
    # Track file descriptions from FLMD files
    variable_file_descriptions = {}  # maps variable -> set of file descriptions

    # Load the file as a polars dataframe
    filetable = pl.read_csv(filetable_path, separator="\t")

    # Create dictionaries for tracking errors
    errors = {
        "failed_urls": [],
        "response_errors": [],
        "encoding_errors": [],
        "unsupported_filetype": []
    }

    # Get the set of entries with an encoding value we can parse
    # This includes CSV files and other parsable formats
    tab_data_files = filetable.filter(
        (pl.col("encoding").is_in(PARSIBLE_ENCODINGS) |
         pl.col("name").str.to_lowercase().str.contains("|".join([ext + "$" for ext in PARSIBLE_EXTENSIONS]))) &
        ~pl.col("name").str.to_lowercase().str.starts_with("readme")
    )

    # Get the set of entries that are XML files (could be EML)
    xml_files = filetable.filter(
        (pl.col("name").str.to_lowercase().str.ends_with(".xml"))
    )

    # Get the set of entries that look like they are data dictionaries
    data_dict_files = filetable.filter(pl.col("name").str.contains("dd.csv$"))

    # Get the set of entries that are FLMD (File Level Metadata) files
    flmd_files = filetable.filter(
        pl.col("name").str.to_lowercase().str.contains("flmd\\.csv$") |
        pl.col("name").str.to_lowercase().str.contains("^flmd\\.csv$")
    )

    # Remove data dict files, XML files, and FLMD files from the tabular files to avoid duplication
    tab_data_files = tab_data_files.join(data_dict_files, on="url", how="anti")
    tab_data_files = tab_data_files.join(xml_files, on="url", how="anti")
    tab_data_files = tab_data_files.join(flmd_files, on="url", how="anti")

    # Initialize the output file
    with open(variable_names_path, "w") as f:
        f.write("name\tfrequency\tsource\tvariable_name\tunits\tdefinition\tdataset\tdataset_name\tfile_description\n")

    # Track variables found in each dataset (discovered during file processing)
    # maps variable -> set of (dataset_id, dataset_name) tuples
    variable_to_datasets = {}

    # First, process FLMD files to build file description mappings
    dataset_file_descriptions = {}  # maps dataset_id -> {filename -> description}
    if len(flmd_files) > 0:
        logging.info(f"Processing {len(flmd_files)} FLMD files...")
        for idx, row in enumerate(tqdm(flmd_files.iter_rows(named=True), desc="Processing FLMD files", unit="file", total=len(flmd_files))):
            url = row["url"]
            filename = row["name"]
            dataset_id = row["dataset_id"]

            try:
                response = requests.get(url, headers=USER_HEADERS, verify=True)
                if response.status_code == 200:
                    try:
                        # Try to decode the content
                        try:
                            response_text = response.content.decode('utf-8')
                        except UnicodeDecodeError:
                            try:
                                response_text = response.content.decode(
                                    'latin-1')
                            except UnicodeDecodeError:
                                response_text = response.content.decode(
                                    'utf-8', errors='ignore')
                                logger.warning(
                                    f"Had to ignore encoding errors for FLMD file {url}")

                        # Parse FLMD file
                        file_desc_mapping = parse_flmd_file(response_text)
                        if file_desc_mapping:
                            if dataset_id not in dataset_file_descriptions:
                                dataset_file_descriptions[dataset_id] = {}
                            dataset_file_descriptions[dataset_id].update(
                                file_desc_mapping)

                    except Exception as e:
                        errors["encoding_errors"].append(f"{url} ({str(e)})")
                        logger.debug(
                            f"Error processing FLMD file {filename}: {str(e)}")
                        continue
                else:
                    errors["response_errors"].append(
                        f"{url} (status code: {response.status_code})")
                    continue

            except Exception as e:
                errors["failed_urls"].append(f"{url} ({str(e)})")
                continue

    # Process files
    file_count = len(tab_data_files) + len(xml_files)
    logging.info(f"Processing {file_count} files...")

    # First process XML files for keywords
    if len(xml_files) > 0:
        for idx, row in enumerate(tqdm(xml_files.iter_rows(named=True), desc="Processing XML files", unit="file", total=len(xml_files))):
            url = row["url"]
            filename = row["name"]
            dataset_id = row["dataset_id"]

            try:
                response = requests.get(url, headers=USER_HEADERS, verify=True)
                status_code = response.status_code

                if (status_code == 200):
                    try:
                        # Try to decode the content as UTF-8
                        try:
                            response_text = response.content.decode('utf-8')
                        except UnicodeDecodeError:
                            # If UTF-8 fails, try another common encoding
                            try:
                                response_text = response.content.decode(
                                    'latin-1')
                            except UnicodeDecodeError:
                                # Last resort, ignore errors
                                response_text = response.content.decode(
                                    'utf-8', errors='ignore')
                                logger.warning(
                                    f"Had to ignore encoding errors for {url}")

                        # Extract keywords from the XML
                        keywords = parse_eml_keywords(response_text)

                        # Normalize keywords
                        # We'll treat them like variables though
                        keywords = normalize_variables(keywords)

                        # Track dataset association for these keywords
                        dataset_name = dataset_id_to_name.get(dataset_id, "")
                        for keyword in keywords:
                            if keyword not in variable_to_datasets:
                                variable_to_datasets[keyword] = set()
                            variable_to_datasets[keyword].add(
                                (dataset_id, dataset_name))

                        # Update keyword frequencies
                        new_keywords = False
                        for keyword in keywords:
                            if keyword in keyword_frequencies:
                                keyword_frequencies[keyword] += 1
                            else:
                                keyword_frequencies[keyword] = 1
                                new_keywords = True

                        # If we found new keywords, append them to the file
                        if new_keywords:
                            with open(variable_names_path, "a") as f:
                                for keyword, freq in keyword_frequencies.items():
                                    if freq == 1:  # Only write new keywords
                                        # Keywords typically don't have units or definitions
                                        # Get dataset info for this keyword (using sets to ensure uniqueness)
                                        dataset_ids = set()
                                        dataset_names = set()
                                        if keyword in dataset_mapping:
                                            for dataset_id, dataset_name in dataset_mapping[keyword]:
                                                dataset_ids.add(dataset_id)
                                                dataset_names.add(dataset_name)
                                        dataset_ids_str = "|".join(
                                            sorted(dataset_ids)) if dataset_ids else ""
                                        dataset_names_str = "|".join(
                                            sorted(dataset_names)) if dataset_names else ""
                                        # Keywords don't have file descriptions
                                        f.write(
                                            f"{keyword}\t{freq}\tkeyword\t{keyword}\t\t\t{dataset_ids_str}\t{dataset_names_str}\t\n")

                    except Exception as e:
                        errors["encoding_errors"].append(f"{url} ({str(e)})")
                        logger.debug(
                            f"Error processing XML file {filename}: {str(e)}")
                        continue
                else:
                    errors["response_errors"].append(
                        f"{url} (status code: {response.status_code})")
                    continue

            except Exception as e:
                errors["failed_urls"].append(f"{url} ({str(e)})")
                continue

    # Now retrieve the column names, then the data dictionaries
    # We will retain the data dictionaries
    total_dd_rows = 0
    for i, fileset_name in enumerate([("data files", tab_data_files), ("data dictionaries", data_dict_files)]):
        name, fileset = fileset_name

        for idx, row in enumerate(tqdm(fileset.iter_rows(named=True), desc=f"Processing {name}", unit="file", total=len(fileset))):
            url = row["url"]
            filename = row["name"]
            dataset_id = row["dataset_id"]
            file_ext = get_file_extension(filename)

            try:
                response = requests.get(
                    url, headers=USER_HEADERS, verify=True, stream=True if file_ext not in [".xls", ".xlsx"] else False
                )
                status_code = response.status_code

                if status_code == 200:
                    try:
                        data_names = []

                        # Process the file based on its type
                        if i == 1 or file_ext == ".csv" or file_ext == ".txt" or file_ext == ".tsv":
                            # Data dictionary or CSV-like file
                            content = response.content

                            # Try to decode the content as UTF-8
                            try:
                                response_text = content.decode('utf-8')
                            except UnicodeDecodeError:
                                # If UTF-8 fails, try another common encoding
                                try:
                                    response_text = content.decode('latin-1')
                                except UnicodeDecodeError:
                                    # Last resort, ignore errors
                                    response_text = content.decode(
                                        'utf-8', errors='ignore')
                                    logger.warning(
                                        f"Had to ignore encoding errors for {url}")

                            if i == 1:  # Data dictionary
                                # Append full DD content to compiled TSV
                                try:
                                    total_dd_rows += append_dd_content_to_file(
                                        response_text,
                                        row["dataset_id"],
                                        filename,
                                        data_dict_compiled_path,
                                    )
                                except Exception as e:
                                    errors["encoding_errors"].append(
                                        f"{url} ({str(e)})")
                                    logger.debug(
                                        f"Error parsing DD content for {filename}: {str(e)}")
                                # Extract names and definitions for addition to the variable names list
                                data_names = parse_data_dictionary(
                                    response_text)
                                dd_definitions, dd_units = parse_dd_defs_units(
                                    response_text)
                                # Update variable definitions, combining multiple definitions with pipes (ensuring uniqueness)
                                for var_name, definition in dd_definitions.items():
                                    if var_name in variable_definitions:
                                        existing_defs = set(
                                            variable_definitions[var_name].split("|"))
                                        existing_defs.add(definition)
                                        variable_definitions[var_name] = "|".join(
                                            sorted(existing_defs))
                                    else:
                                        variable_definitions[var_name] = definition
                                # Update variable units, combining multiple units with pipes (ensuring uniqueness)
                                for var_name, unit in dd_units.items():
                                    if var_name in variable_units:
                                        existing_units = set(
                                            variable_units[var_name].split("|"))
                                        existing_units.add(unit)
                                        variable_units[var_name] = "|".join(
                                            sorted(existing_units))
                                    else:
                                        variable_units[var_name] = unit
                            else:  # CSV file
                                # For CSV, just get the header
                                for line in response_text.splitlines():
                                    if line.strip():  # Skip empty lines
                                        data_names = parse_header(line)
                                        break  # We only need the first line

                        elif file_ext == ".xlsx" or file_ext == ".xls":
                            # Excel file
                            content = response.content
                            data_names = parse_excel_header(content, filename)

                        else:
                            errors["unsupported_filetype"].append(
                                f"{url} (filetype: {file_ext})")
                            continue

                        # Normalize the column names
                        data_names = normalize_variables(data_names)

                        # Track dataset association for these variables
                        dataset_name = dataset_id_to_name.get(dataset_id, "")
                        for var_name in data_names:
                            if var_name not in variable_to_datasets:
                                variable_to_datasets[var_name] = set()
                            variable_to_datasets[var_name].add(
                                (dataset_id, dataset_name))

                        # Get file description for this file if available
                        current_dataset_id = row["dataset_id"]
                        current_filename = filename
                        file_description = ""
                        if (current_dataset_id in dataset_file_descriptions and
                                current_filename in dataset_file_descriptions[current_dataset_id]):
                            file_description = dataset_file_descriptions[current_dataset_id][current_filename]

                        # Update file descriptions for variables from this file
                        if file_description:
                            for var_name in data_names:
                                if var_name not in variable_file_descriptions:
                                    variable_file_descriptions[var_name] = set(
                                    )
                                variable_file_descriptions[var_name].add(
                                    file_description)

                        # Update frequencies based on source type
                        if i == 1:  # Data dictionary
                            new_data_dict_vars = False
                            for name in data_names:
                                if name in data_dict_frequencies:
                                    data_dict_frequencies[name] += 1
                                else:
                                    data_dict_frequencies[name] = 1
                                    new_data_dict_vars = True

                            # If we found new data dictionary variables, append them to the file
                            if new_data_dict_vars:
                                with open(variable_names_path, "a") as f:
                                    for dd_var, freq in data_dict_frequencies.items():
                                        if freq == 1:  # Only write new variables
                                            var_name, _ = extract_units(dd_var)
                                            definition = variable_definitions.get(
                                                dd_var, "")
                                            units = variable_units.get(
                                                dd_var, "")
                                            # Get dataset info for this variable (using sets to ensure uniqueness)
                                            dataset_ids = set()
                                            dataset_names = set()
                                            if dd_var in dataset_mapping:
                                                for dataset_id, dataset_name in dataset_mapping[dd_var]:
                                                    dataset_ids.add(dataset_id)
                                                    dataset_names.add(
                                                        dataset_name)
                                            dataset_ids_str = "|".join(
                                                sorted(dataset_ids)) if dataset_ids else ""
                                            dataset_names_str = "|".join(
                                                sorted(dataset_names)) if dataset_names else ""
                                            # Get file descriptions for this variable
                                            file_desc_set = variable_file_descriptions.get(
                                                dd_var, set())
                                            file_desc_str = "|".join(
                                                sorted(file_desc_set)) if file_desc_set else ""
                                            f.write(
                                                f"{dd_var}\t{freq}\tdata_dictionary\t{var_name}\t{units}\t{definition}\t{dataset_ids_str}\t{dataset_names_str}\t{file_desc_str}\n")
                        else:  # Regular data files
                            new_columns = False
                            for name in data_names:
                                if name in column_frequencies:
                                    column_frequencies[name] += 1
                                else:
                                    column_frequencies[name] = 1
                                    new_columns = True

                            # If we found new variables, append them to the file
                            if new_columns:
                                with open(variable_names_path, "a") as f:
                                    for column, freq in column_frequencies.items():
                                        if freq == 1:  # Only write new variables
                                            var_name, unit = extract_units(
                                                column)
                                            # Columns don't have definitions or units from data dictionaries
                                            # Get dataset info for this column (using sets to ensure uniqueness)
                                            dataset_ids = set()
                                            dataset_names = set()
                                            if column in dataset_mapping:
                                                for dataset_id, dataset_name in dataset_mapping[column]:
                                                    dataset_ids.add(dataset_id)
                                                    dataset_names.add(
                                                        dataset_name)
                                            dataset_ids_str = "|".join(
                                                sorted(dataset_ids)) if dataset_ids else ""
                                            dataset_names_str = "|".join(
                                                sorted(dataset_names)) if dataset_names else ""
                                            # Get file descriptions for this column
                                            file_desc_set = variable_file_descriptions.get(
                                                column, set())
                                            file_desc_str = "|".join(
                                                sorted(file_desc_set)) if file_desc_set else ""
                                            f.write(
                                                f"{column}\t{freq}\tcolumn\t{var_name}\t{unit}\t\t{dataset_ids_str}\t{dataset_names_str}\t{file_desc_str}\n")

                    except Exception as e:
                        errors["encoding_errors"].append(f"{url} ({str(e)})")
                        logger.debug(f"Encoding error for {url}: {str(e)}")
                        continue
                else:
                    errors["response_errors"].append(
                        f"{url} (status code: {response.status_code})")
                    continue

            except Exception as e:
                errors["failed_urls"].append(f"{url} ({str(e)})")
                continue

    # After processing all files, update the file with final frequencies
    # First identify terms that appear in multiple sources and combine their sources
    # Create a dictionary to track combined term frequencies and sources
    all_terms = {}

    # Get all unique terms across all sources
    all_unique_terms = set()
    all_unique_terms.update(column_frequencies.keys())
    all_unique_terms.update(keyword_frequencies.keys())
    all_unique_terms.update(data_dict_frequencies.keys())

    # For each unique term, determine its sources and combined frequency
    for term in all_unique_terms:
        sources = []
        total_freq = 0

        if term in column_frequencies:
            sources.append("column")
            total_freq += column_frequencies[term]

        if term in keyword_frequencies:
            sources.append("keyword")
            total_freq += keyword_frequencies[term]

        if term in data_dict_frequencies:
            sources.append("data_dictionary")
            total_freq += data_dict_frequencies[term]

        # Combine sources with pipe delimiter
        combined_source = "|".join(sources)
        all_terms[term] = (total_freq, combined_source)

    # Sort by frequency (highest first)
    sorted_terms = sorted(
        all_terms.items(), key=lambda item: item[1][0], reverse=True)

    with open(variable_names_path, "w") as f:
        f.write("name\tfrequency\tsource\tvariable_name\tunits\tdefinition\tdataset\tdataset_name\tfile_description\n")
        for term, (frequency, source) in sorted_terms:
            var_name, _ = extract_units(term)
            definition = variable_definitions.get(term, "")
            units = variable_units.get(term, "")

            # Get dataset IDs and names for this variable (using sets to ensure uniqueness)
            dataset_ids = set()
            dataset_names = set()

            # Include datasets from the results file mapping (pre-existing variables)
            if term in dataset_mapping:
                for dataset_id, dataset_name in dataset_mapping[term]:
                    dataset_ids.add(dataset_id)
                    dataset_names.add(dataset_name)

            # Include datasets discovered during file processing
            if term in variable_to_datasets:
                for dataset_id, dataset_name in variable_to_datasets[term]:
                    dataset_ids.add(dataset_id)
                    dataset_names.add(dataset_name)

            # Join with pipes for multiple datasets (sorted for consistency)
            dataset_ids_str = "|".join(
                sorted(dataset_ids)) if dataset_ids else ""
            dataset_names_str = "|".join(
                sorted(dataset_names)) if dataset_names else ""

            # Get file descriptions for this variable
            file_desc_set = variable_file_descriptions.get(term, set())
            file_desc_str = "|".join(
                sorted(file_desc_set)) if file_desc_set else ""

            f.write(
                f"{term}\t{frequency}\t{source}\t{var_name}\t{units}\t{definition}\t{dataset_ids_str}\t{dataset_names_str}\t{file_desc_str}\n")

    # Log any errors that occurred during processing
    if total_dd_rows:
        logger.info(
            f"Compiled {total_dd_rows} data dictionary rows into {data_dict_compiled_path}")
    if errors["response_errors"]:
        logger.error(
            f"Response errors for {len(errors['response_errors'])} URLs")
        for url in errors["response_errors"][:5]:  # Log first few errors
            logger.error(f"  {url}")
        if len(errors["response_errors"]) > 5:
            logger.error(f"  ...and {len(errors['response_errors']) - 5} more")

    if errors["failed_urls"]:
        logger.error(f"Failed to process {len(errors['failed_urls'])} URLs")
        for url in errors["failed_urls"][:5]:  # Log first few errors
            logger.error(f"  {url}")
        if len(errors["failed_urls"]) > 5:
            logger.error(f"  ...and {len(errors['failed_urls']) - 5} more")

    if errors["encoding_errors"]:
        logger.error(
            f"Encoding errors for {len(errors['encoding_errors'])} URLs")
        for url in errors["encoding_errors"][:5]:  # Log first few errors
            logger.error(f"  {url}")
        if len(errors["encoding_errors"]) > 5:
            logger.error(f"  ...and {len(errors['encoding_errors']) - 5} more")

    if errors["unsupported_filetype"]:
        logger.error(
            f"Unsupported file types for {len(errors['unsupported_filetype'])} URLs")
        for url in errors["unsupported_filetype"][:5]:  # Log first few errors
            logger.error(f"  {url}")
        if len(errors["unsupported_filetype"]) > 5:
            logger.error(
                f"  ...and {len(errors['unsupported_filetype']) - 5} more")

    return variable_names_path


def normalize_variables(variables: list) -> list:
    """Normalize variables from ESS-DIVE.

    This function:
    1. Splits hierarchical terms (those with '>')
    2. Removes Unicode special characters (BOM, zero-width spaces, etc.)
    3. Removes leading and trailing punctuation (preserves parentheses and brackets)
    4. Converts to lowercase
    5. Replaces underscores with spaces
    6. Strips whitespace
    7. Filters out purely numeric names (e.g., "1234" or "123.45")
    8. Sorts the list
    """
    normalized = []
    for var in variables:
        if not var:  # Skip empty strings
            continue

        # Convert to string if it's not already
        var = str(var)

        # Clean Unicode special characters
        var = clean_unicode_chars(var)

        if ">" in var:
            # This is a hierarchy but we just want all terms
            vars_split = var.split(">")
            for v in vars_split:
                if not v.strip():  # Skip empty strings after splitting
                    continue
                v = v.lower().replace("_", " ").strip()

                # Clean leading and trailing punctuation while preserving parentheses and brackets
                v = clean_punctuation(v, preserve_brackets=True)

                # Apply all filtering logic
                if _is_valid_variable_name(v) and v not in normalized and len(v) <= 70:
                    normalized.append(v)
        else:
            var = var.lower().replace("_", " ").strip()

            # Clean leading and trailing punctuation while preserving parentheses and brackets
            var = clean_punctuation(var, preserve_brackets=True)

            # Apply all filtering logic
            if _is_valid_variable_name(var) and var not in normalized and len(var) <= 70:
                normalized.append(var)

    normalized.sort()
    return normalized


def _is_valid_variable_name(name: str) -> bool:
    """Check if a variable name is valid (not purely numeric, not empty, etc.)."""
    if not name or not name.strip():
        return False

    clean_name = name.strip()

    # Remove leading punctuation and non-alphanumeric characters
    while clean_name and (clean_name[0] in string.punctuation or not clean_name[0].isalnum()):
        clean_name = clean_name[1:]

    if not clean_name:
        return False

    # Skip specific metadata terms (case-insensitive)
    excluded_terms = {"n/a", "notes", "file_name", "header_rows", "data_type"}
    if clean_name.lower().replace("_", " ").replace(" ", "_") in excluded_terms or \
       clean_name.lower().replace("_", " ") in excluded_terms or \
       clean_name.lower().replace(" ", "_") in excluded_terms:
        return False

    # Skip purely numeric names (integers or decimals)
    try:
        float(clean_name.replace(" ", ""))  # Remove spaces for number check
        # If conversion succeeds, it's purely numeric - reject it
        return False
    except ValueError:
        # If conversion fails, it contains non-numeric characters - keep it
        return True


def parse_header(header: str) -> list:
    """Parse header from a data file.
    Also normalizes and filters out excessively long column names (>70 chars)."""
    header_names = []

    # Clean Unicode special characters first
    header = clean_unicode_chars(header)

    reader = csv.reader(StringIO(header))
    for row in reader:
        for name in row:
            # Skip empty names and names > 70 chars
            if name != "" and len(name) <= 70:
                clean_name = clean_unicode_chars(
                    name.lower().replace("_", " ").strip())

                # Remove leading punctuation and non-alphanumeric characters
                while clean_name and (clean_name[0] in string.punctuation or not clean_name[0].isalnum()):
                    clean_name = clean_name[1:]

                if clean_name:  # Only add if not empty after cleaning
                    header_names.append(clean_name)
    return header_names


def parse_data_dictionary(dd: str) -> list:
    """Parse a data dictionary.
    Extracts variable names from the first column and applies normalize_variables
    for consistent filtering and cleaning."""
    data_names = []

    # Clean Unicode special characters first
    dd = clean_unicode_chars(dd)

    reader = csv.reader(StringIO(dd))
    for row in reader:
        if not row:  # Skip empty rows
            continue
        name = row[0]
        # Skip headers and collect all non-empty names
        if name not in ["", "Column_or_Row_Name"]:
            data_names.append(name)

    # Apply consistent normalization (this will handle all filtering)
    return normalize_variables(data_names)


def parse_dd_defs_units(dd: str) -> tuple:
    """Parse a data dictionary and extract variable names with their definitions and units.
    Returns:
        Tuple of two dicts: (definitions, units)
        Each dict maps normalized variable names to their values (pipe-delimited if multiple).
    """
    var_definitions = {}
    var_units = {}
    dd = clean_unicode_chars(dd)
    reader = csv.DictReader(StringIO(dd))
    if not reader.fieldnames:
        return var_definitions, var_units

    # Find the definition and unit columns (case-insensitive)
    definition_col = None
    unit_col = None
    for field in reader.fieldnames:
        norm_field = _norm_header_key(field)
        if norm_field == "definition":
            definition_col = field
        if norm_field == "unit":
            unit_col = field
    if not definition_col and not unit_col:
        return var_definitions, var_units
    for row in reader:
        if not row:
            continue

        # First column is variable name
        name = row.get(reader.fieldnames[0], "")
        definition = row.get(definition_col, "") if definition_col else ""
        unit = row.get(unit_col, "") if unit_col else ""
        if name and name not in ["", "Column_or_Row_Name"]:
            normalized_names = normalize_variables([name])
            if normalized_names:
                normalized_name = normalized_names[0]

                # Get definitions
                if definition and definition.strip():
                    def_val = sanitize_tsv_field(definition.strip())
                    if normalized_name in var_definitions:
                        existing_defs = var_definitions[normalized_name].split(
                            "|")
                        if def_val not in existing_defs:
                            var_definitions[normalized_name] += f"|{def_val}"
                    else:
                        var_definitions[normalized_name] = def_val

                # Get units
                if unit and unit.strip():
                    unit_val = sanitize_tsv_field(unit.strip())
                    if normalized_name in var_units:
                        existing_units = var_units[normalized_name].split("|")
                        if unit_val not in existing_units:
                            var_units[normalized_name] += f"|{unit_val}"
                    else:
                        var_units[normalized_name] = unit_val

    return var_definitions, var_units


def parse_eml_keywords(content: str) -> List[str]:
    """Parse keywords from an Ecological Metadata Language (EML) XML file.
    Also filters out excessively long keywords (>70 chars).

    Args:
        content: String containing the XML content

    Returns:
        List of keywords extracted from the XML
    """
    keywords = []

    try:
        # Parse the XML content
        # The namespace is important for finding elements correctly
        root = ET.fromstring(content)

        # Find all keyword elements, regardless of their namespace path
        # This handles different EML structures and namespaces
        for keyword_elem in root.findall(".//*{https://eml.ecoinformatics.org/eml-2.2.0}keyword") or \
                root.findall(".//*keyword") or \
                root.findall(".//*{*}keyword"):
            if keyword_elem.text and keyword_elem.text.strip():
                keyword = keyword_elem.text.strip().lower()
                if len(keyword) <= 70:  # Skip keywords > 70 chars
                    keywords.append(keyword)
    except Exception as e:
        logger.debug(f"Error parsing EML XML: {str(e)}")

    return keywords
